import openpyxl
import Levenshtein

workbook = openpyxl.load_workbook("Тестовое_задание_аналитика_данных.xlsx")
sheet_list = workbook["Список"]

# функция для решения 3й задачи
def check_similar_brands(brand_to_check, all_brands):
    similar_brands = []
    for brand in all_brands:
        distance = Levenshtein.distance(str(brand_to_check).lower(), str(brand).lower())
        if distance <= 2:
            similar_brands.append(brand)
    return similar_brands

# функция для обращения к столбцу по его названию
def get_ind(sheet, column_name):
    for idx, cell in enumerate(sheet[1]):
        if cell.value == column_name:
            return idx+1
    return None

#Задача 1. Названия брендов большими буквами без пробелов.
brand_ = get_ind(sheet_list, 'brand_')
brand = get_ind(sheet_list, 'brand')

for i, row in enumerate(sheet_list.iter_rows(min_row=2, min_col=brand_, max_col=brand_, values_only=True), start=2):
    old_brand = row[0]
    new_brand = old_brand.upper().strip()
    sheet_list.cell(row=i, column=brand, value=new_brand)

# Задача 2. Если наименование бренда пустое, или написано
# "отсутствует"/"без товарного знака" и т.п., то необходимо
# найти данный товар в интернете и проставить значение Бренда

# Сделала просто вывод сообщения об ошибке
print("Необходимо заполнить наименования брендов:")
for row in range(2, sheet_list.max_row + 1):
    brand_name = sheet_list.cell(row=row, column=brand_).value
    if not brand_name or any(word in brand_name.lower() for word in ["отсутствует", "без товарного знака"]):
        product_id = sheet_list.cell(row=row, column=get_ind(sheet_list, 'gtin_')).value
        print(f"gtin товара: {product_id}, строка: {row}")


#Задача 3. Необходимо проверить похожие написания Брендов.
# Возможно, поставщик мог ошибиться на 1 букву или поставить
# лишний знак и 1 бренд физически превратился в 2

# Считаю с помощью функции нахождения расстояния Левенштейна

all_brands = []
if brand_:
    for cell in sheet_list[brand_][1:]:
        if cell.value is not None:
            all_brands.append(cell.value)

# проверка каждого бренда на похожие написания
for row in sheet_list.iter_rows(min_row=2, min_col=brand_, max_col=brand_):
    brand_to_check = row[0].value
    similar_brands = check_similar_brands(brand_to_check, all_brands)

# Задача 4.Необходимо по ТОП 200 Брендов проверить корректность заполнения данного параметра
brand_top_dict = {}
top_brand = get_ind(sheet_list, 'ТОП Бренд')

# Проверяю, что у каждого бренда всегда одинаковое числовое значение в столбце ТОП Бренд,
# Если у бренда меняется значение в колонке Топ Бренд, то исправляю на его предыдущее значение в ТОПе
for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    brand_value = row[brand]
    top_brand_value = row[top_brand]

    if brand_value not in brand_top_dict:
        brand_top_dict[brand] = top_brand_value
    else:
        if top_brand_value != brand_top_dict[brand_value]:
            sheet_list.cell(row=i, column=top_brand, value=brand_top_dict[brand])


# Задача 5.Необходимо в новое поле "group" указать параметр "Йогурт питьевой" или "Йогурт ложковой"
package_type_ = get_ind(sheet_list, 'package_type_')-1
product_type_ = get_ind(sheet_list, 'product_type_')-1
package = get_ind(sheet_list, 'package')

# определяю по столбцу product_type или по типу упаковки
for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    package_type = row[package_type_]
    product_type = row[product_type_]

    if product_type == 'ЙОГУРТ ПИТЬЕВОЙ' or package_type in ['КУВШИН', 'БУТЫЛКА', 'ПАКЕТ ПРЯМОУГОЛЬНЫЙ',
                                                             'ПАКЕТ БЕЗ ФОРМЫ', 'ВЕДРО']:
        group = 'йогурт питьевой'
    else:
        group = 'Йогурт ложковой'

    sheet_list.cell(row=i, column=get_ind(sheet_list, 'group'), value=group)


# Задача 6. для ТОП 200 необходимо проверить параметр package_type_
# и занести корретное значение в поле "package".


for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    package_type = row[package_type_]
    if package_type == 'БАНКА НЕМЕТАЛЛИЧЕСКАЯ':
        package_value = 'ст/банка'
    elif package_type in ['ПАКЕТ БЕЗ ФОРМЫ', 'ПАКЕТ ПРЯМОУГОЛЬНЫЙ']:
        package_value = 'пакет'
    elif package_type == 'КОРОБКА/БОКС':
        package_value = 'петрапак'
    elif package_type == 'НЕТ В СПРАВОЧНИКЕ':
        package_value = None
    elif package_type == 'СТАКАН':
            package_value = 'пл/стакан'
    else:
        package_value = str.lower(package_type)
    sheet_list.cell(row=i, column=package, value=package_value)

# Задача 7. для ТОП 200 необходимо проверить значение fat_
# и привести в единый вид все наименования жирностей

fat_ = get_ind(sheet_list, 'fat_')-1
fat = get_ind(sheet_list, 'fat')

for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    fat_value = row[fat_].replace('%', '')
    formatted_fat_value = '{:.1f}%'.format(float(fat_value.replace(',', '.')))
    sheet_list.cell(row=i, column=fat, value=formatted_fat_value)


# Задача 8. необходимо все веса привести в грамм/мл.
# новые значение вынести в поле "weight". Например 0,26 кг - указать - 260

volume_weight_ = get_ind(sheet_list, 'volume_weight_')-1
weight_measure_ = get_ind(sheet_list, 'weight_measure_unit_')-1
weight = get_ind(sheet_list, 'weight')-1

for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    volume_value = row[volume_weight_]
    weight_value = row[weight_measure_]
    if weight_value.lower() == 'кг' or weight_value.lower() == 'л':
        weight_in_grams = volume_value * 1000
    else:
        weight_in_grams = volume_value
    sheet_list.cell(row=i, column=weight+1, value=weight_in_grams)

# Задача 9. Необходимо вынести вкус в отдельную колонку.
# Примеры написания вкусов на листе "taste". Бифидобактерии, на кокосовмо молоке и т.п. - не является вкусом!

# сначала удаляю из листа taste строки с неподходящими словами:
sheet_taste = workbook["taste"]

words_to_delete = ["натуральный", "бифидо", "био", "ароматизированный", "в ассортименте", "кисломолочный",
                   "козий", "овечий", "из", "без сахара", "с сахаром", "несладкий", "сладкий",
                   "нежный", "томленый", "специи", "иммунолакт", "термостатный",
                   "безлактозный", "коллаген", "мюсли", "с наполнителем", "фруктовый", "сливочный", "сахар",
                   "персика", "детский"]

rows_to_delete = []

for row in sheet_taste.iter_rows(min_row=1, max_row=sheet_taste.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    if any(word in cell_value for word in words_to_delete):
        rows_to_delete.append(row[0].row)

for row_number in rows_to_delete:
    sheet_taste.cell(row=row_number, column=1, value="")

taste = get_ind(sheet_list, 'taste')
product_name_ = get_ind(sheet_list, 'product_name_')-1

# создаю список вкусов из листа taste, в котором уже удалены неподходящие слова

keywords = []
for row in sheet_taste.iter_rows(min_row=2, max_row=sheet_taste.max_row, min_col=1, max_col=1, values_only=True):
    cell_value = row[0]
    if cell_value:
        words = cell_value.split(",")
        keywords.extend([word.strip() for word in words])

for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    description = str(row[product_name_])
    taste_words = []  # создаю список для слов в стобце product_name_, которые содержат слова из keyweords - списка вкусов
    for keyword in keywords:
        if keyword.lower() in description.lower():
            taste_words.append(keyword)

    if taste_words:
        sheet_list.cell(row=i, column=taste, value=', '.join(taste_words))
    else:
        sheet_list.cell(row=i, column=taste, value="без вкуса")


# Задача 10. В новое поле base записываются параметры по примеру во вкладке "base"

sheet_base = workbook["base"]

# cоздаю словарь баз для йогурта из листа base
base_dict = {}
for row in sheet_base.iter_rows(values_only=True):
    base_dict[row[0]] = row[0]
base_header = 'base'
sheet_list.cell(row=1, column=sheet_list.max_column + 1, value=base_header)

# если в описании встречаются слова из словаря, то в поле base заношу эту базу.
# Иначе заношу просто product type
for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    description = str(row[product_name_])
    base_value = row[product_type_]
    for base_item in base_dict.values():
        if base_item in description:
            base_value = base_item
            break
    sheet_list.cell(row=i, column=sheet_list.max_column, value=str.lower(base_value))

# Задача 11
# новое поле sku. Необходимо собрать новое наименование товара
# по принципу сцепки параметров bra+base+taste+fat+weight+package -
# пример новых наименований на вкладке "sku"

sku = get_ind(sheet_list, 'sku')
base = get_ind(sheet_list, 'base')-1

for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    brand_s = str(row[brand-1]).strip()
    base_s = str(row[base]).strip()
    taste_s = str(row[taste-1]).strip()
    fat_s = str(row[fat-1]).strip()
    weight_s = str(row[weight]).strip()
    package_s = str(row[package-1]).strip()

    new_name = f"{brand_s} {base_s} {taste_s} м.д.ж. {fat_s} {weight_s}г {package_s}"
    sheet_list.cell(row=i, column=sku, value=new_name)

# Задача 12. параметр срока годности товара. Бывают fresh/uht (длительный срок).
# Данный параметр также необходимо проставить для ТОП 200

# Если в поле storage_condition указан срок хранения 35 дней,
# то записываю значение uht - длительного хранения. Иначе - fresh
storage_condition_ = get_ind(sheet_list, 'storage_condition_')-1

expiration_header = 'expiration'
sheet_list.cell(row=1, column=sheet_list.max_column+1, value=expiration_header)

for i, row in enumerate(sheet_list.iter_rows(min_row=2, values_only=True), start=2):
    expiration = ''
    storage_condition = str(row[storage_condition_]).strip()
    if '"type":"1"' in storage_condition:
        expiration = 'uht'
    else:
        expiration = 'fresh'
    sheet_list.cell(row=i, column=sheet_list.max_column, value=expiration)

workbook.save("Test.xlsx")
